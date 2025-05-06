#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Simülasyon ve Strateji Modülü
"""

import os
import time
import json
import logging
from datetime import datetime, timedelta
import pandas as pd
import numpy as np
import threading
import queue
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

import config

logger = logging.getLogger(__name__)


class TradingSimulator:
    """Alım-satım simülasyonu sınıfı"""

    def __init__(self, api_manager, model_manager, data_fetcher):
        """
        Simülatörü başlatır

        Args:
            api_manager: API yöneticisi
            model_manager: Model yöneticisi
            data_fetcher: Veri çekme sınıfı
        """
        self.api_manager = api_manager
        self.model_manager = model_manager
        self.data_fetcher = data_fetcher

        # Simülasyon parametreleri
        self.initial_balance = config.DEFAULT_SIMULATION_PARAMS['initial_balance']
        self.fee_percent = config.DEFAULT_SIMULATION_PARAMS['fee_percent'] / 100.0
        self.stop_loss_percent = config.DEFAULT_SIMULATION_PARAMS['stop_loss_percent'] / 100.0
        self.take_profit_percent = config.DEFAULT_SIMULATION_PARAMS['take_profit_percent'] / 100.0
        self.max_open_positions = config.DEFAULT_SIMULATION_PARAMS['max_open_positions']
        self.position_size_percent = config.DEFAULT_SIMULATION_PARAMS['position_size_percent'] / 100.0

        # Simülasyon durumu
        self.results = []
        self.positions = []
        self.balance = self.initial_balance
        self.equity = self.initial_balance
        self.is_running = False
        self.start_time = None
        self.end_time = None
        self.real_time_mode = False
        self.result_callback = None
        self.log_callback = None
        self.status_callback = None
        self.progress_callback = None
        self.message_queue = queue.Queue()

        # Klasörleri oluştur
        os.makedirs(config.PATHS['results_dir'], exist_ok=True)

    def set_parameters(self, **kwargs):
        """
        Simülasyon parametrelerini ayarlar

        Args:
            **kwargs: Parametre anahtar-değer çiftleri
        """
        if 'initial_balance' in kwargs:
            self.initial_balance = float(kwargs['initial_balance'])
            self.balance = self.initial_balance
            self.equity = self.initial_balance

        if 'fee_percent' in kwargs:
            self.fee_percent = float(kwargs['fee_percent']) / 100.0

        if 'stop_loss_percent' in kwargs:
            self.stop_loss_percent = float(kwargs['stop_loss_percent']) / 100.0

        if 'take_profit_percent' in kwargs:
            self.take_profit_percent = float(kwargs['take_profit_percent']) / 100.0

        if 'max_open_positions' in kwargs:
            self.max_open_positions = int(kwargs['max_open_positions'])

        if 'position_size_percent' in kwargs:
            self.position_size_percent = float(kwargs['position_size_percent']) / 100.0

        logger.info(f"Simülasyon parametreleri güncellendi: Bakiye={self.initial_balance:.2f}, "
                    f"İşlem ücreti={self.fee_percent * 100:.2f}%, Stop Loss={self.stop_loss_percent * 100:.2f}%, "
                    f"Take Profit={self.take_profit_percent * 100:.2f}%, Pozisyon Oranı={self.position_size_percent * 100:.2f}%")

    def run_simulation(self, data, model_path=None, real_time=False, progress_callback=None, **params):
        """
        Simülasyonu başlatır (backtest veya gerçek zamanlı).
        """
        self.progress_callback = progress_callback
        self.set_parameters(**params)


        if real_time:
            symbol = data
            interval = model_path
            self.is_running = True
            self.real_time_mode = True
            loaded = self.model_manager.load_model(model_path)
            if not loaded:
                self.log_callback(f"❌ Model yüklenemedi: {model_path}", error=True)
                return
            else:
                self.log_callback(f"✅ Model yüklendi: {model_path}")

            if self.log_callback:
                self.log_callback("[Gerçek Zamanlı] Simülasyon başlatıldı.")

            try:
                while self.is_running:
                    try:
                        df = self.data_fetcher.fetch_latest_data(symbol, interval, limit=100)

                        if df is None or not isinstance(df, pd.DataFrame) or df.empty:
                            error_msg = "Geçersiz veri alındı" if df is not None else "Veri alınamadı (None döndü)"
                            self.log_callback(f"[HATA] {error_msg}. 10 saniye bekleniyor...", True)
                            time.sleep(10)
                            continue

                        lookback = self.model_manager.lookback
                        features = self.model_manager.feature_columns

                        if len(df) < lookback:
                            self.log_callback(f"Yeterli veri yok ({len(df)}/{lookback}). Bekleniyor...", True)
                            time.sleep(10)
                            continue

                        x = np.array([df[features].values[-lookback:]])

                        predicted = self.model_manager.predict_next(x)
                        if predicted is None:
                            self.log_callback("Tahmin başarısız oldu: predicted=None", error=True)
                            time.sleep(10)
                            continue

                        current_price = df["close"].iloc[-1]
                        if current_price is None:
                            self.log_callback("Mevcut fiyat verisi alınamadı", error=True)
                            time.sleep(10)
                            continue

                        direction = np.sign(predicted - current_price)

                        log_msg = (f"[{datetime.now().strftime('%H:%M:%S')}] "
                                   f"Fiyat: {current_price:.4f} | "
                                   f"Tahmin: {predicted:.4f} | "
                                   f"Yön: {'↑' if direction > 0 else '↓' if direction < 0 else '→'}")

                        if self.log_callback:
                            self.log_callback(log_msg)

                        if self.progress_callback:
                            try:
                                self.progress_callback(0, 1, log_msg)
                            except Exception as e:
                                self.log_callback(f"Callback hatası: {str(e)}", True)

                        # İşlem büyüklüğü ve ücret hesapla
                        if direction != 0 and self.balance > 0:
                            quantity = self.balance * self.position_size_percent / current_price
                            fee = self.balance * self.fee_percent
                            pnl = quantity * current_price * (self.take_profit_percent if direction > 0 else -self.stop_loss_percent)

                            self.balance += pnl - fee
                            self.equity = self.balance

                            self.positions.append({
                                'entry_time': datetime.now(),
                                'entry_price': current_price,
                                'exit_time': datetime.now(),
                                'exit_price': current_price * (1 + self.take_profit_percent if direction > 0 else 1 - self.stop_loss_percent),
                                'pnl': pnl,
                                'fee': fee,
                                'size': quantity,
                                'status': 'closed',
                                'type': 'long' if direction > 0 else 'short',
                                'symbol': symbol,
                                'exit_reason': 'TP/SL',
                                'pnl_percent': pnl / (self.balance + 1e-9) * 100,
                                'id': len(self.positions) + 1
                            })

                        time.sleep(300)

                    except Exception as e:
                        self.log_callback(f"İşlem hatası: {str(e)}", True)
                        time.sleep(30)
                        continue

            except KeyboardInterrupt:
                self.log_callback("Kullanıcı tarafından durduruldu", False)
            except Exception as e:
                self.log_callback(f"Kritik hata: {str(e)}. Simülasyon durduruluyor!", True)
            finally:
                self.is_running = False
                self.log_callback("Simülasyon durduruldu", False)
            return None

            return None
        else:
            # Backtest kodu burada kalacak
            pass

    def _manage_realtime_positions(self, current_price, direction):
        """Gerçek zamanlı pozisyon yönetimi"""
        try:
            # Açık pozisyonları kontrol et
            for position in self.positions:
                if position['status'] == 'open':
                    entry_price = position['entry_price']

                    if position['type'] == 'long':
                        pnl_pct = (current_price - entry_price) / entry_price
                    else:  # short
                        pnl_pct = (entry_price - current_price) / entry_price

                    # Stop loss/take profit kontrolü
                    if pnl_pct <= -self.stop_loss_percent:
                        self.close_position(position, datetime.now(), current_price, 'stop_loss')
                    elif pnl_pct >= self.take_profit_percent:
                        self.close_position(position, datetime.now(), current_price, 'take_profit')

            # Yeni pozisyon açma kontrolü
            if direction != 0 and self.balance > 0:
                open_positions = sum(1 for p in self.positions if p['status'] == 'open')
                if open_positions < self.max_open_positions:
                    position_type = 'long' if direction > 0 else 'short'
                    position_size = self.balance * self.position_size_percent

                    self.open_position(
                        symbol=self.model_manager.symbol,
                        position_type=position_type,
                        entry_time=datetime.now(),
                        entry_price=current_price,
                        size=position_size
                    )

        except Exception as e:
            self.log_callback(f"Pozisyon yönetimi hatası: {str(e)}", True)
    def stop_simulation(self):
        """Simülasyonu durdurur"""
        if self.is_running:
            self.is_running = False
            self.log("Simülasyon durduruldu.")
            return True
        return False

        # ... (Diğer metodların geri kalanı)

    def calculate_statistics(self):
            """
            Simülasyon istatistiklerini hesaplar

            Returns:
                dict: İstatistikler
            """
            if not self.positions:
                return {
                    'total_trades': 0,
                    'profitable_trades': 0,
                    'win_rate': 0,
                    'avg_profit': 0,
                    'avg_loss': 0,
                    'total_pnl': 0,
                    'total_pnl_percent': 0,
                    'max_drawdown': 0,
                    'max_drawdown_percent': 0,
                    'sharpe_ratio': 0
                }

            # Temel istatistikler
            total_trades = len([p for p in self.positions if p['status'] == 'closed'])
            profitable_trades = len([p for p in self.positions if p['status'] == 'closed' and p['pnl'] > 0])
            loss_trades = len([p for p in self.positions if p['status'] == 'closed' and p['pnl'] <= 0])

            win_rate = profitable_trades / total_trades if total_trades > 0 else 0

            # Kar/zarar istatistikleri
            profits = [p['pnl'] for p in self.positions if p['status'] == 'closed' and p['pnl'] > 0]
            losses = [p['pnl'] for p in self.positions if p['status'] == 'closed' and p['pnl'] <= 0]

            avg_profit = np.mean(profits) if profits else 0
            avg_loss = np.mean(losses) if losses else 0

            total_pnl = sum(p['pnl'] for p in self.positions if p['status'] == 'closed')
            total_pnl_percent = (self.balance / self.initial_balance - 1) * 100

            # Maksimum drawdown
            balances = [self.initial_balance]
            for p in sorted(self.positions,
                            key=lambda x: x['exit_time'] if x['exit_time'] is not None else datetime.max):
                if p['status'] == 'closed':
                    balances.append(balances[-1] + p['pnl'])

            cummax = np.maximum.accumulate(balances)
            drawdowns = (cummax - balances) / cummax * 100
            max_drawdown_percent = np.max(drawdowns) if len(drawdowns) > 0 else 0
            max_drawdown = np.max(cummax - balances) if len(balances) > 0 else 0

            # Sharpe oranı
            daily_returns = []

            if len(self.results) > 1:
                daily_returns = pd.Series([r['balance'] for r in self.results]).pct_change().dropna().values

            sharpe_ratio = np.mean(daily_returns) / np.std(daily_returns) * np.sqrt(252) if len(
                daily_returns) > 0 and np.std(daily_returns) > 0 else 0

            # Sonuçları döndür
            return {
                'total_trades': total_trades,
                'profitable_trades': profitable_trades,
                'loss_trades': loss_trades,
                'win_rate': win_rate,
                'avg_profit': avg_profit,
                'avg_loss': avg_loss,
                'avg_profit_loss_ratio': abs(avg_profit / avg_loss) if avg_loss != 0 else 0,
                'total_pnl': total_pnl,
                'total_pnl_percent': total_pnl_percent,
                'max_drawdown': max_drawdown,
                'max_drawdown_percent': max_drawdown_percent,
                'sharpe_ratio': sharpe_ratio
            }

    def export_results_to_excel(self, results_df, positions_df=None, filename=None):
            """
            Sonuçları Excel'e aktarır

            Args:
                results_df (pd.DataFrame): Sonuç verileri
                positions_df (pd.DataFrame): Pozisyon verileri
                filename (str): Dosya adı

            Returns:
                str: Oluşturulan dosya yolu
            """
            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{self.model_manager.symbol}_{self.model_manager.interval}_sim_{timestamp}.xlsx"

            filepath = os.path.join(config.PATHS['results_dir'], filename)

            # Excel dosyası oluştur
            wb = Workbook()

            # Sonuçlar sayfası
            ws_results = wb.active
            ws_results.title = config.EXCEL_SETTINGS['sheet_names']['sonuclar']

            # Başlıkları ekle
            headers = ['Zaman', 'Fiyat', 'Tahmin', 'Tahmin Yönü', 'Gerçek Yön', 'Doğru mu?', 'Bakiye', 'Varlık']
            for col_idx, header in enumerate(headers, 1):
                cell = ws_results.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Verileri ekle
            for row_idx, row in enumerate(results_df.to_dict('records'), 2):
                ws_results.cell(row=row_idx, column=1).value = row.get('timestamp', '').strftime(
                    '%Y-%m-%d %H:%M:%S') if isinstance(row.get('timestamp'), datetime) else row.get('timestamp', '')
                ws_results.cell(row=row_idx, column=2).value = row.get('close', 0)
                ws_results.cell(row=row_idx, column=3).value = row.get('predicted_close', 0)
                ws_results.cell(row=row_idx, column=4).value = row.get('predicted_direction', 0)
                ws_results.cell(row=row_idx, column=5).value = row.get('actual_direction', 0)
                ws_results.cell(row=row_idx, column=6).value = 'Evet' if row.get('is_correct', False) else 'Hayır'
                ws_results.cell(row=row_idx, column=7).value = row.get('balance', 0)
                ws_results.cell(row=row_idx, column=8).value = row.get('equity', 0)

            # Sütun genişliklerini ayarla
            for col_idx in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_idx)
                ws_results.column_dimensions[col_letter].width = 15

            # Pozisyonlar sayfası
            if positions_df is not None and not positions_df.empty:
                ws_positions = wb.create_sheet(title="İşlemler")

                # Başlıkları ekle
                headers = ['ID', 'Sembol', 'Tip', 'Durum', 'Giriş Zamanı', 'Giriş Fiyatı', 'Miktar', 'Ücret',
                           'Çıkış Zamanı', 'Çıkış Fiyatı', 'Çıkış Nedeni', 'Kar/Zarar', 'Kar/Zarar %']

                for col_idx, header in enumerate(headers, 1):
                    cell = ws_positions.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')

                # Verileri ekle
                for row_idx, row in enumerate(positions_df.to_dict('records'), 2):
                    ws_positions.cell(row=row_idx, column=1).value = row.get('id', 0)
                    ws_positions.cell(row=row_idx, column=2).value = row.get('symbol', '')
                    ws_positions.cell(row=row_idx, column=3).value = row.get('type', '').upper()
                    ws_positions.cell(row=row_idx, column=4).value = 'AÇIK' if row.get('status',
                                                                                       '') == 'open' else 'KAPALI'
                    ws_positions.cell(row=row_idx, column=5).value = row.get('entry_time', '').strftime(
                        '%Y-%m-%d %H:%M:%S') if isinstance(row.get('entry_time'), datetime) else row.get('entry_time',
                                                                                                         '')
                    ws_positions.cell(row=row_idx, column=6).value = row.get('entry_price', 0)
                    ws_positions.cell(row=row_idx, column=7).value = row.get('size', 0)
                    ws_positions.cell(row=row_idx, column=8).value = row.get('fee', 0)
                    ws_positions.cell(row=row_idx, column=9).value = row.get('exit_time', '').strftime(
                        '%Y-%m-%d %H:%M:%S') if isinstance(row.get('exit_time'), datetime) else row.get('exit_time', '')
                    ws_positions.cell(row=row_idx, column=10).value = row.get('exit_price', 0)
                    ws_positions.cell(row=row_idx, column=11).value = row.get('exit_reason', '')
                    ws_positions.cell(row=row_idx, column=12).value = row.get('pnl', 0)
                    ws_positions.cell(row=row_idx, column=13).value = row.get('pnl_percent', 0)

                    # Karlı/zararlı işlemleri renklendir
                    if row.get('status', '') == 'closed':
                        pnl = row.get('pnl', 0)
                        pnl_cell = ws_positions.cell(row=row_idx, column=12)

                        if pnl > 0:
                            pnl_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                                                        fill_type="solid")  # Yeşil
                        elif pnl < 0:
                            pnl_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                                        fill_type="solid")  # Kırmızı

                # Sütun genişliklerini ayarla
                for col_idx in range(1, len(headers) + 1):
                    col_letter = get_column_letter(col_idx)
                    ws_positions.column_dimensions[col_letter].width = 15

            # Özet sayfası
            ws_summary = wb.create_sheet(title=config.EXCEL_SETTINGS['sheet_names']['ozet'])

            summary_stats = self.calculate_statistics()

            # Başlık
            ws_summary.cell(row=1, column=1).value = "Simülasyon Özeti"
            ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
            ws_summary.merge_cells('A1:C1')

            # Temel bilgiler
            basic_info = [
                ("Sembol", self.model_manager.symbol),
                ("Aralık", self.model_manager.interval),
                ("Başlangıç Bakiyesi", f"${self.initial_balance:.2f}"),
                ("Son Bakiye", f"${self.balance:.2f}"),
                ("Toplam Kar/Zarar",
                 f"${self.balance - self.initial_balance:.2f} ({(self.balance / self.initial_balance - 1) * 100:.2f}%)"),
                ("Toplam İşlem Sayısı", summary_stats['total_trades']),
                ("Karlı İşlemler", f"{summary_stats['profitable_trades']} ({summary_stats['win_rate'] * 100:.2f}%)"),
                ("Zararlı İşlemler", summary_stats['loss_trades']),
                ("Ortalama Kar", f"${summary_stats['avg_profit']:.2f}"),
                ("Ortalama Zarar", f"${abs(summary_stats['avg_loss']):.2f}"),
                ("Kar/Zarar Oranı", f"{summary_stats['avg_profit_loss_ratio']:.2f}"),
                ("Maksimum Drawdown",
                 f"${summary_stats['max_drawdown']:.2f} ({summary_stats['max_drawdown_percent']:.2f}%)"),
                ("Sharpe Oranı", f"{summary_stats['sharpe_ratio']:.2f}"),
                ("Simülasyon Süresi", f"{(self.end_time - self.start_time):.2f} saniye")
            ]

            # Özet bilgileri ekle
            for row_idx, (label, value) in enumerate(basic_info, 3):
                ws_summary.cell(row=row_idx, column=1).value = label
                ws_summary.cell(row=row_idx, column=1).font = Font(bold=True)
                ws_summary.cell(row=row_idx, column=2).value = value

            # Sütun genişliklerini ayarla
            ws_summary.column_dimensions['A'].width = 25
            ws_summary.column_dimensions['B'].width = 25

            # Parametre sayfası
            ws_params = wb.create_sheet(title=config.EXCEL_SETTINGS['sheet_names']['ayarlar'])

            # Başlık
            ws_params.cell(row=1, column=1).value = "Simülasyon Parametreleri"
            ws_params.cell(row=1, column=1).font = Font(bold=True, size=14)
            ws_params.merge_cells('A1:C1')

            # Parametre bilgileri
            params_info = [
                ("Başlangıç Bakiyesi", f"${self.initial_balance:.2f}"),
                ("İşlem Ücreti", f"{self.fee_percent * 100:.2f}%"),
                ("Stop Loss", f"{self.stop_loss_percent * 100:.2f}%"),
                ("Take Profit", f"{self.take_profit_percent * 100:.2f}%"),
                ("Maksimum Açık Pozisyon", self.max_open_positions),
                ("Pozisyon Boyutu", f"{self.position_size_percent * 100:.2f}%")
            ]

            # Model parametreleri
            model_params = [
                ("Model Türü", "LSTM"),
                ("Lookback Periyodu", self.model_manager.lookback),
                ("Özellik Sayısı", len(self.model_manager.feature_columns)),
                ("Kullanılan Özellikler", ", ".join(self.model_manager.feature_columns)),
                ("Eğitim/Test Oranı",
                 f"{self.model_manager.train_size * 100:.0f}/{(1 - self.model_manager.train_size) * 100:.0f}")
            ]

            # Parametre bilgilerini ekle
            for row_idx, (label, value) in enumerate(params_info, 3):
                ws_params.cell(row=row_idx, column=1).value = label
                ws_params.cell(row=row_idx, column=1).font = Font(bold=True)
                ws_params.cell(row=row_idx, column=2).value = value

            # Model bilgilerini ekle
            ws_params.cell(row=len(params_info) + 4, column=1).value = "Model Parametreleri"
            ws_params.cell(row=len(params_info) + 4, column=1).font = Font(bold=True, size=12)

            for row_idx, (label, value) in enumerate(model_params, len(params_info) + 6):
                ws_params.cell(row=row_idx, column=1).value = label
                ws_params.cell(row=row_idx, column=1).font = Font(bold=True)
                ws_params.cell(row=row_idx, column=2).value = value

            # Sütun genişliklerini ayarla
            ws_params.column_dimensions['A'].width = 25
            ws_params.column_dimensions['B'].width = 40

            # Dosyayı kaydet
            wb.save(filepath)

            self.log(f"Sonuçlar Excel dosyasına kaydedildi: {filepath}")
            return filepath

    def create_performance_chart(self, results_df, show_predictions=True):
            """
            Performans grafiği oluşturur

            Args:
                results_df (pd.DataFrame): Sonuç verileri
                show_predictions (bool): Tahminler gösterilsin mi

            Returns:
                fig: Matplotlib figürü
            """
            # Figür oluştur
            fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 8), gridspec_kw={'height_ratios': [3, 1]})

            # Üst grafik: Fiyat ve tahminler
            if show_predictions:
                ax1.plot(results_df.index, results_df['close'], label='Gerçek Fiyat', color='blue')
                ax1.plot(results_df.index, results_df['predicted_close'], label='Tahmin', color='red', linestyle='--')
                ax1.set_title(f'{self.model_manager.symbol} {self.model_manager.interval} - Fiyat ve Tahminler')
            else:
                ax1.plot(results_df.index, results_df['close'], label='Fiyat', color='blue')
                ax1.set_title(f'{self.model_manager.symbol} {self.model_manager.interval} - Fiyat Grafiği')

            ax1.set_ylabel('Fiyat')
            ax1.legend(loc='upper left')
            ax1.grid(True)

            # Alt grafik: Bakiye değişimi
            ax2.plot(results_df.index, results_df['balance'], label='Bakiye', color='green')
            ax2.set_xlabel('Zaman')
            ax2.set_ylabel('Bakiye ($)')
            ax2.legend(loc='upper left')
            ax2.grid(True)

            # Pozisyonları ekle
            closed_positions = [p for p in self.positions if p['status'] == 'closed']

            for pos in closed_positions:
                entry_time = pos['entry_time']
                exit_time = pos['exit_time']

                if entry_time in results_df.index and exit_time in results_df.index:
                    entry_idx = results_df.index.get_loc(entry_time)
                    exit_idx = results_df.index.get_loc(exit_time)

                    entry_price = pos['entry_price']
                    exit_price = pos['exit_price']

                    # Giriş ve çıkış noktalarını işaretle
                    if pos['type'] == 'long':
                        ax1.scatter(entry_time, entry_price, marker='^', color='green', s=100)
                        ax1.scatter(exit_time, exit_price, marker='v', color='red', s=100)
                    else:  # short
                        ax1.scatter(entry_time, entry_price, marker='v', color='red', s=100)
                        ax1.scatter(exit_time, exit_price, marker='^', color='green', s=100)

            plt.tight_layout()

            return fig


